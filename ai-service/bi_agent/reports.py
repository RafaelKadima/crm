"""
Report Generator - Gerador de Relatórios
=========================================

Responsável por gerar relatórios em PDF e Excel.
Usa WeasyPrint para PDF e OpenPyXL para Excel.
"""

import logging
import os
import io
from datetime import datetime
from typing import Optional, Dict, Any, List
from pathlib import Path
import base64

logger = logging.getLogger(__name__)

# Tenta importar bibliotecas opcionais
try:
    from weasyprint import HTML, CSS
    WEASYPRINT_AVAILABLE = True
except ImportError:
    WEASYPRINT_AVAILABLE = False
    logger.warning("WeasyPrint não disponível. Instale com: pip install weasyprint")

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logger.warning("OpenPyXL não disponível. Instale com: pip install openpyxl")


class ReportGenerator:
    """
    Gerador de relatórios do BI Agent.
    
    Formatos suportados:
    - PDF (usando WeasyPrint)
    - Excel (usando OpenPyXL)
    """
    
    def __init__(self, tenant_id: str):
        self.tenant_id = tenant_id
        self.reports_dir = Path("/tmp/bi_reports")
        self.reports_dir.mkdir(exist_ok=True)
    
    def generate_pdf(
        self,
        title: str,
        data: Dict[str, Any],
        report_type: str = "executive_summary"
    ) -> Dict[str, Any]:
        """
        Gera relatório em PDF.
        
        Args:
            title: Título do relatório
            data: Dados para popular o relatório
            report_type: Tipo de relatório (executive_summary, sales, marketing, etc.)
        
        Returns:
            Dict com status, path ou base64 do PDF
        """
        if not WEASYPRINT_AVAILABLE:
            return {
                "success": False,
                "error": "WeasyPrint não instalado. Adicione ao requirements.txt.",
            }
        
        try:
            # Gera HTML baseado no tipo
            html_content = self._generate_html(title, data, report_type)
            
            # Converte para PDF
            pdf_bytes = self._html_to_pdf(html_content)
            
            # Salva ou retorna base64
            filename = f"{report_type}_{self.tenant_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
            filepath = self.reports_dir / filename
            
            with open(filepath, "wb") as f:
                f.write(pdf_bytes)
            
            # Retorna também em base64 para download direto
            pdf_base64 = base64.b64encode(pdf_bytes).decode("utf-8")
            
            return {
                "success": True,
                "filename": filename,
                "path": str(filepath),
                "base64": pdf_base64,
                "size_bytes": len(pdf_bytes),
            }
            
        except Exception as e:
            logger.error(f"[ReportGenerator] Erro ao gerar PDF: {e}")
            return {
                "success": False,
                "error": str(e),
            }
    
    def _generate_html(self, title: str, data: Dict, report_type: str) -> str:
        """Gera HTML para o relatório."""
        if report_type == "executive_summary":
            return self._html_executive_summary(title, data)
        elif report_type == "sales":
            return self._html_sales_report(title, data)
        elif report_type == "marketing":
            return self._html_marketing_report(title, data)
        else:
            return self._html_generic_report(title, data)
    
    def _html_executive_summary(self, title: str, data: Dict) -> str:
        """Template HTML para resumo executivo."""
        revenue = data.get("revenue", {})
        leads = data.get("leads", {})
        roas = data.get("roas", {})
        highlights = data.get("highlights", [])
        
        return f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="utf-8">
            <style>
                body {{
                    font-family: 'Segoe UI', Tahoma, sans-serif;
                    margin: 40px;
                    color: #333;
                    line-height: 1.6;
                }}
                .header {{
                    background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                    color: white;
                    padding: 30px;
                    margin: -40px -40px 30px -40px;
                    border-radius: 0 0 10px 10px;
                }}
                .header h1 {{
                    margin: 0;
                    font-size: 28px;
                }}
                .header .date {{
                    opacity: 0.8;
                    font-size: 14px;
                }}
                .kpi-grid {{
                    display: grid;
                    grid-template-columns: repeat(2, 1fr);
                    gap: 20px;
                    margin-bottom: 30px;
                }}
                .kpi-card {{
                    background: #f8f9fa;
                    border-radius: 10px;
                    padding: 20px;
                    border-left: 4px solid #667eea;
                }}
                .kpi-card h3 {{
                    margin: 0 0 10px 0;
                    font-size: 14px;
                    color: #666;
                    text-transform: uppercase;
                }}
                .kpi-card .value {{
                    font-size: 32px;
                    font-weight: bold;
                    color: #333;
                }}
                .kpi-card .change {{
                    font-size: 14px;
                    color: #28a745;
                }}
                .kpi-card .change.negative {{
                    color: #dc3545;
                }}
                .highlights {{
                    background: #fff3cd;
                    border: 1px solid #ffc107;
                    border-radius: 10px;
                    padding: 20px;
                    margin-bottom: 30px;
                }}
                .highlights h2 {{
                    margin: 0 0 15px 0;
                    color: #856404;
                }}
                .highlights ul {{
                    margin: 0;
                    padding-left: 20px;
                }}
                .footer {{
                    margin-top: 40px;
                    text-align: center;
                    color: #999;
                    font-size: 12px;
                }}
            </style>
        </head>
        <body>
            <div class="header">
                <h1>{title}</h1>
                <div class="date">Gerado em {datetime.now().strftime('%d/%m/%Y às %H:%M')}</div>
            </div>
            
            <div class="kpi-grid">
                <div class="kpi-card">
                    <h3>Receita</h3>
                    <div class="value">R$ {revenue.get('current', 0):,.2f}</div>
                    <div class="change {'negative' if revenue.get('growth', 0) < 0 else ''}">
                        {revenue.get('growth', 0)*100:+.1f}% vs período anterior
                    </div>
                </div>
                
                <div class="kpi-card">
                    <h3>Leads</h3>
                    <div class="value">{leads.get('total', 0)}</div>
                    <div class="change">
                        {leads.get('rate', 0)*100:.1f}% taxa de conversão
                    </div>
                </div>
                
                <div class="kpi-card">
                    <h3>ROAS</h3>
                    <div class="value">{roas.get('current', 0):.2f}x</div>
                    <div class="change">
                        Melhor: {roas.get('best_campaign', 'N/A')}
                    </div>
                </div>
                
                <div class="kpi-card">
                    <h3>Conversões</h3>
                    <div class="value">{leads.get('converted', 0)}</div>
                    <div class="change">
                        de {leads.get('total', 0)} leads
                    </div>
                </div>
            </div>
            
            {'<div class="highlights"><h2>🔑 Destaques</h2><ul>' + ''.join(f'<li>{h}</li>' for h in highlights) + '</ul></div>' if highlights else ''}
            
            <div class="footer">
                Relatório gerado automaticamente pelo BI Agent
            </div>
        </body>
        </html>
        """
    
    def _html_sales_report(self, title: str, data: Dict) -> str:
        """Template HTML para relatório de vendas."""
        return f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="utf-8">
            <style>
                body {{ font-family: 'Segoe UI', sans-serif; margin: 40px; }}
                .header {{ background: #2e7d32; color: white; padding: 20px; margin: -40px -40px 30px; }}
                table {{ width: 100%; border-collapse: collapse; margin: 20px 0; }}
                th, td {{ padding: 12px; text-align: left; border-bottom: 1px solid #ddd; }}
                th {{ background: #f5f5f5; }}
            </style>
        </head>
        <body>
            <div class="header">
                <h1>{title}</h1>
                <p>{datetime.now().strftime('%d/%m/%Y')}</p>
            </div>
            
            <h2>Resumo de Vendas</h2>
            <table>
                <tr><th>Métrica</th><th>Valor</th></tr>
                <tr><td>Total de Leads</td><td>{data.get('total_leads', 0)}</td></tr>
                <tr><td>Leads no Período</td><td>{data.get('leads_in_period', 0)}</td></tr>
                <tr><td>Taxa de Conversão</td><td>{data.get('conversion_rate', 0)*100:.1f}%</td></tr>
                <tr><td>Valor Total</td><td>R$ {data.get('total_value', 0):,.2f}</td></tr>
                <tr><td>Ticket Médio</td><td>R$ {data.get('avg_deal_size', 0):,.2f}</td></tr>
            </table>
        </body>
        </html>
        """
    
    def _html_marketing_report(self, title: str, data: Dict) -> str:
        """Template HTML para relatório de marketing."""
        campaigns = data.get("campaigns", [])
        
        campaigns_rows = ""
        for c in campaigns[:10]:
            campaigns_rows += f"""
            <tr>
                <td>{c.get('name', 'N/A')}</td>
                <td>{c.get('status', 'N/A')}</td>
                <td>R$ {c.get('spend', 0):,.2f}</td>
                <td>{c.get('clicks', 0)}</td>
                <td>{c.get('roas', 0):.2f}x</td>
            </tr>
            """
        
        return f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="utf-8">
            <style>
                body {{ font-family: 'Segoe UI', sans-serif; margin: 40px; }}
                .header {{ background: #1976d2; color: white; padding: 20px; margin: -40px -40px 30px; }}
                table {{ width: 100%; border-collapse: collapse; margin: 20px 0; }}
                th, td {{ padding: 12px; text-align: left; border-bottom: 1px solid #ddd; }}
                th {{ background: #f5f5f5; }}
                .kpi {{ display: inline-block; padding: 15px 25px; background: #e3f2fd; margin: 10px; border-radius: 8px; }}
            </style>
        </head>
        <body>
            <div class="header">
                <h1>{title}</h1>
                <p>{datetime.now().strftime('%d/%m/%Y')}</p>
            </div>
            
            <div class="kpis">
                <div class="kpi">
                    <strong>Gasto Total</strong><br>
                    R$ {data.get('total_spend', 0):,.2f}
                </div>
                <div class="kpi">
                    <strong>ROAS Médio</strong><br>
                    {data.get('roas', 0):.2f}x
                </div>
                <div class="kpi">
                    <strong>CPL</strong><br>
                    R$ {data.get('cpl', 0):.2f}
                </div>
            </div>
            
            <h2>Campanhas</h2>
            <table>
                <tr>
                    <th>Nome</th>
                    <th>Status</th>
                    <th>Gasto</th>
                    <th>Cliques</th>
                    <th>ROAS</th>
                </tr>
                {campaigns_rows if campaigns_rows else '<tr><td colspan="5">Nenhuma campanha encontrada</td></tr>'}
            </table>
        </body>
        </html>
        """
    
    def _html_generic_report(self, title: str, data: Dict) -> str:
        """Template HTML genérico."""
        rows = ""
        for key, value in data.items():
            if isinstance(value, dict):
                rows += f"<tr><td colspan='2'><strong>{key}</strong></td></tr>"
                for k, v in value.items():
                    rows += f"<tr><td style='padding-left:20px'>{k}</td><td>{v}</td></tr>"
            else:
                rows += f"<tr><td>{key}</td><td>{value}</td></tr>"
        
        return f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="utf-8">
            <style>
                body {{ font-family: 'Segoe UI', sans-serif; margin: 40px; }}
                .header {{ background: #333; color: white; padding: 20px; margin: -40px -40px 30px; }}
                table {{ width: 100%; border-collapse: collapse; }}
                td {{ padding: 10px; border-bottom: 1px solid #eee; }}
            </style>
        </head>
        <body>
            <div class="header">
                <h1>{title}</h1>
                <p>{datetime.now().strftime('%d/%m/%Y às %H:%M')}</p>
            </div>
            <table>{rows}</table>
        </body>
        </html>
        """
    
    def _html_to_pdf(self, html_content: str) -> bytes:
        """Converte HTML para PDF usando WeasyPrint."""
        html = HTML(string=html_content)
        pdf_bytes = html.write_pdf()
        return pdf_bytes
    
    def generate_excel(
        self,
        title: str,
        data: Dict[str, Any],
        report_type: str = "executive_summary"
    ) -> Dict[str, Any]:
        """
        Gera relatório em Excel.
        
        Args:
            title: Título do relatório
            data: Dados para popular o relatório
            report_type: Tipo de relatório
        
        Returns:
            Dict com status, path ou base64 do arquivo
        """
        if not OPENPYXL_AVAILABLE:
            return {
                "success": False,
                "error": "OpenPyXL não instalado. Adicione ao requirements.txt.",
            }
        
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Relatório"
            
            # Estilos
            header_font = Font(bold=True, size=14, color="FFFFFF")
            header_fill = PatternFill("solid", fgColor="667eea")
            subheader_font = Font(bold=True, size=11)
            border = Border(
                bottom=Side(style='thin', color='DDDDDD')
            )
            
            # Cabeçalho
            ws.merge_cells('A1:D1')
            ws['A1'] = title
            ws['A1'].font = Font(bold=True, size=16)
            
            ws.merge_cells('A2:D2')
            ws['A2'] = f"Gerado em {datetime.now().strftime('%d/%m/%Y às %H:%M')}"
            ws['A2'].font = Font(italic=True, color="666666")
            
            current_row = 4
            
            # Adiciona dados baseado no tipo
            if report_type == "executive_summary":
                current_row = self._excel_executive_summary(ws, data, current_row, header_fill, header_font)
            elif report_type == "sales":
                current_row = self._excel_sales_data(ws, data, current_row, header_fill, header_font)
            elif report_type == "marketing":
                current_row = self._excel_marketing_data(ws, data, current_row, header_fill, header_font)
            else:
                current_row = self._excel_generic_data(ws, data, current_row)
            
            # Ajusta largura das colunas
            for col in range(1, 6):
                ws.column_dimensions[get_column_letter(col)].width = 20
            
            # Salva
            filename = f"{report_type}_{self.tenant_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            filepath = self.reports_dir / filename
            wb.save(filepath)
            
            # Lê para base64
            with open(filepath, "rb") as f:
                excel_bytes = f.read()
            excel_base64 = base64.b64encode(excel_bytes).decode("utf-8")
            
            return {
                "success": True,
                "filename": filename,
                "path": str(filepath),
                "base64": excel_base64,
                "size_bytes": len(excel_bytes),
            }
            
        except Exception as e:
            logger.error(f"[ReportGenerator] Erro ao gerar Excel: {e}")
            return {
                "success": False,
                "error": str(e),
            }
    
    def _excel_executive_summary(self, ws, data: Dict, row: int, header_fill, header_font) -> int:
        """Adiciona dados de resumo executivo ao Excel."""
        # Seção: Receita
        ws[f'A{row}'] = "Receita"
        ws[f'A{row}'].fill = header_fill
        ws[f'A{row}'].font = header_font
        ws.merge_cells(f'A{row}:D{row}')
        row += 1
        
        revenue = data.get("revenue", {})
        ws[f'A{row}'] = "Atual"
        ws[f'B{row}'] = f"R$ {revenue.get('current', 0):,.2f}"
        row += 1
        ws[f'A{row}'] = "Anterior"
        ws[f'B{row}'] = f"R$ {revenue.get('previous', 0):,.2f}"
        row += 1
        ws[f'A{row}'] = "Crescimento"
        ws[f'B{row}'] = f"{revenue.get('growth', 0)*100:+.1f}%"
        row += 2
        
        # Seção: Leads
        ws[f'A{row}'] = "Leads"
        ws[f'A{row}'].fill = header_fill
        ws[f'A{row}'].font = header_font
        ws.merge_cells(f'A{row}:D{row}')
        row += 1
        
        leads = data.get("leads", {})
        ws[f'A{row}'] = "Total"
        ws[f'B{row}'] = leads.get("total", 0)
        row += 1
        ws[f'A{row}'] = "Convertidos"
        ws[f'B{row}'] = leads.get("converted", 0)
        row += 1
        ws[f'A{row}'] = "Taxa de Conversão"
        ws[f'B{row}'] = f"{leads.get('rate', 0)*100:.1f}%"
        row += 2
        
        # Destaques
        highlights = data.get("highlights", [])
        if highlights:
            ws[f'A{row}'] = "Destaques"
            ws[f'A{row}'].fill = header_fill
            ws[f'A{row}'].font = header_font
            ws.merge_cells(f'A{row}:D{row}')
            row += 1
            
            for h in highlights:
                ws[f'A{row}'] = f"• {h}"
                ws.merge_cells(f'A{row}:D{row}')
                row += 1
        
        return row
    
    def _excel_sales_data(self, ws, data: Dict, row: int, header_fill, header_font) -> int:
        """Adiciona dados de vendas ao Excel."""
        ws[f'A{row}'] = "Métricas de Vendas"
        ws[f'A{row}'].fill = header_fill
        ws[f'A{row}'].font = header_font
        ws.merge_cells(f'A{row}:D{row}')
        row += 1
        
        metrics = [
            ("Total de Leads", data.get("total_leads", 0)),
            ("Leads no Período", data.get("leads_in_period", 0)),
            ("Taxa de Conversão", f"{data.get('conversion_rate', 0)*100:.1f}%"),
            ("Valor Total", f"R$ {data.get('total_value', 0):,.2f}"),
            ("Ticket Médio", f"R$ {data.get('avg_deal_size', 0):,.2f}"),
        ]
        
        for label, value in metrics:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            row += 1
        
        return row
    
    def _excel_marketing_data(self, ws, data: Dict, row: int, header_fill, header_font) -> int:
        """Adiciona dados de marketing ao Excel."""
        # KPIs
        ws[f'A{row}'] = "KPIs de Marketing"
        ws[f'A{row}'].fill = header_fill
        ws[f'A{row}'].font = header_font
        ws.merge_cells(f'A{row}:D{row}')
        row += 1
        
        ws[f'A{row}'] = "Gasto Total"
        ws[f'B{row}'] = f"R$ {data.get('total_spend', 0):,.2f}"
        row += 1
        ws[f'A{row}'] = "ROAS Médio"
        ws[f'B{row}'] = f"{data.get('roas', 0):.2f}x"
        row += 1
        ws[f'A{row}'] = "CPL"
        ws[f'B{row}'] = f"R$ {data.get('cpl', 0):.2f}"
        row += 2
        
        # Campanhas
        campaigns = data.get("campaigns", [])
        if campaigns:
            ws[f'A{row}'] = "Campanhas"
            ws[f'A{row}'].fill = header_fill
            ws[f'A{row}'].font = header_font
            ws.merge_cells(f'A{row}:E{row}')
            row += 1
            
            # Headers
            headers = ["Nome", "Status", "Gasto", "Cliques", "ROAS"]
            for i, h in enumerate(headers, 1):
                ws[f'{get_column_letter(i)}{row}'] = h
                ws[f'{get_column_letter(i)}{row}'].font = Font(bold=True)
            row += 1
            
            for c in campaigns[:20]:
                ws[f'A{row}'] = c.get("name", "N/A")
                ws[f'B{row}'] = c.get("status", "N/A")
                ws[f'C{row}'] = f"R$ {c.get('spend', 0):,.2f}"
                ws[f'D{row}'] = c.get("clicks", 0)
                ws[f'E{row}'] = f"{c.get('roas', 0):.2f}x"
                row += 1
        
        return row
    
    def _excel_generic_data(self, ws, data: Dict, row: int) -> int:
        """Adiciona dados genéricos ao Excel."""
        for key, value in data.items():
            if isinstance(value, dict):
                ws[f'A{row}'] = key
                ws[f'A{row}'].font = Font(bold=True)
                row += 1
                for k, v in value.items():
                    ws[f'B{row}'] = k
                    ws[f'C{row}'] = str(v)
                    row += 1
            else:
                ws[f'A{row}'] = key
                ws[f'B{row}'] = str(value)
                row += 1
        
        return row
    
    def generate_report(
        self,
        title: str,
        data: Dict[str, Any],
        report_type: str = "executive_summary",
        formats: List[str] = ["pdf"]
    ) -> Dict[str, Any]:
        """
        Gera relatório em múltiplos formatos.
        
        Args:
            title: Título do relatório
            data: Dados do relatório
            report_type: Tipo do relatório
            formats: Lista de formatos desejados ["pdf", "excel"]
        
        Returns:
            Dict com resultados de cada formato
        """
        results = {}
        
        if "pdf" in formats:
            results["pdf"] = self.generate_pdf(title, data, report_type)
        
        if "excel" in formats:
            results["excel"] = self.generate_excel(title, data, report_type)
        
        return {
            "success": all(r.get("success", False) for r in results.values()),
            "formats": results,
            "generated_at": datetime.now().isoformat(),
        }

