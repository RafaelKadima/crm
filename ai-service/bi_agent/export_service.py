"""
Export Service - Serviço de Exportação de Relatórios
=====================================================

Gera relatórios em PDF e Excel.
"""

import logging
from datetime import datetime
from typing import Optional, Dict, Any, List
import io
import os

logger = logging.getLogger(__name__)


class PDFExportService:
    """
    Serviço de exportação para PDF.
    
    Usa WeasyPrint para gerar PDFs a partir de HTML.
    """
    
    def __init__(self, tenant_id: str):
        self.tenant_id = tenant_id
        self.output_dir = f"/tmp/reports/{tenant_id}"
        os.makedirs(self.output_dir, exist_ok=True)
    
    async def generate_executive_pdf(
        self,
        data: Dict[str, Any],
        period: str
    ) -> Dict[str, Any]:
        """Gera PDF do relatório executivo."""
        filename = f"executive_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        filepath = os.path.join(self.output_dir, filename)
        
        html = self._render_executive_html(data, period)
        await self._html_to_pdf(html, filepath)
        
        return {
            "filename": filename,
            "filepath": filepath,
            "url": f"/reports/{self.tenant_id}/{filename}",
            "size_kb": os.path.getsize(filepath) // 1024 if os.path.exists(filepath) else 0,
            "generated_at": datetime.now().isoformat(),
        }
    
    async def generate_sales_pdf(
        self,
        data: Dict[str, Any],
        period: str
    ) -> Dict[str, Any]:
        """Gera PDF do relatório de vendas."""
        filename = f"sales_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        filepath = os.path.join(self.output_dir, filename)
        
        html = self._render_sales_html(data, period)
        await self._html_to_pdf(html, filepath)
        
        return {
            "filename": filename,
            "filepath": filepath,
            "url": f"/reports/{self.tenant_id}/{filename}",
            "size_kb": os.path.getsize(filepath) // 1024 if os.path.exists(filepath) else 0,
            "generated_at": datetime.now().isoformat(),
        }
    
    async def generate_marketing_pdf(
        self,
        data: Dict[str, Any],
        period: str
    ) -> Dict[str, Any]:
        """Gera PDF do relatório de marketing."""
        filename = f"marketing_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        filepath = os.path.join(self.output_dir, filename)
        
        html = self._render_marketing_html(data, period)
        await self._html_to_pdf(html, filepath)
        
        return {
            "filename": filename,
            "filepath": filepath,
            "url": f"/reports/{self.tenant_id}/{filename}",
            "size_kb": os.path.getsize(filepath) // 1024 if os.path.exists(filepath) else 0,
            "generated_at": datetime.now().isoformat(),
        }
    
    def _render_executive_html(self, data: Dict, period: str) -> str:
        """Renderiza HTML para relatório executivo."""
        kpis = data.get('kpis', {})
        highlights = data.get('highlights', [])
        
        return f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="UTF-8">
            <title>Relatório Executivo</title>
            <style>
                body {{ font-family: 'Segoe UI', Arial, sans-serif; padding: 40px; color: #333; }}
                h1 {{ color: #1a1a2e; border-bottom: 3px solid #4361ee; padding-bottom: 10px; }}
                .header {{ display: flex; justify-content: space-between; align-items: center; margin-bottom: 30px; }}
                .period {{ background: #f8f9fa; padding: 10px 20px; border-radius: 8px; }}
                .kpis {{ display: grid; grid-template-columns: repeat(3, 1fr); gap: 20px; margin: 30px 0; }}
                .kpi {{ background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); color: white; padding: 20px; border-radius: 12px; }}
                .kpi-label {{ font-size: 14px; opacity: 0.9; }}
                .kpi-value {{ font-size: 32px; font-weight: bold; margin-top: 8px; }}
                .section {{ margin: 30px 0; }}
                .section h2 {{ color: #4361ee; font-size: 18px; }}
                .highlight {{ background: #e8f5e9; padding: 12px 16px; border-radius: 8px; margin: 8px 0; border-left: 4px solid #4caf50; }}
                .alert {{ background: #fff3e0; border-left-color: #ff9800; }}
                .footer {{ margin-top: 40px; padding-top: 20px; border-top: 1px solid #eee; font-size: 12px; color: #666; }}
            </style>
        </head>
        <body>
            <div class="header">
                <h1>Relatório Executivo</h1>
                <div class="period">Período: {period}</div>
            </div>
            
            <div class="kpis">
                <div class="kpi">
                    <div class="kpi-label">Receita</div>
                    <div class="kpi-value">R$ {kpis.get('revenue', {}).get('current', 0):,.0f}</div>
                </div>
                <div class="kpi">
                    <div class="kpi-label">Leads Convertidos</div>
                    <div class="kpi-value">{kpis.get('leads', {}).get('converted', 0)}</div>
                </div>
                <div class="kpi">
                    <div class="kpi-label">Taxa de Conversão</div>
                    <div class="kpi-value">{kpis.get('leads', {}).get('rate', 0) * 100:.1f}%</div>
                </div>
            </div>
            
            <div class="section">
                <h2>Destaques</h2>
                {''.join(f'<div class="highlight">{h}</div>' for h in highlights)}
            </div>
            
            <div class="footer">
                <p>Gerado em {datetime.now().strftime('%d/%m/%Y às %H:%M')} pelo BI Agent</p>
            </div>
        </body>
        </html>
        """
    
    def _render_sales_html(self, data: Dict, period: str) -> str:
        """Renderiza HTML para relatório de vendas."""
        summary = data.get('summary', {})
        funnel = data.get('funnel', {})
        
        return f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="UTF-8">
            <title>Relatório de Vendas</title>
            <style>
                body {{ font-family: 'Segoe UI', Arial, sans-serif; padding: 40px; color: #333; }}
                h1 {{ color: #1a1a2e; border-bottom: 3px solid #10b981; padding-bottom: 10px; }}
                .summary {{ display: grid; grid-template-columns: repeat(4, 1fr); gap: 15px; margin: 30px 0; }}
                .summary-card {{ background: #f8f9fa; padding: 15px; border-radius: 8px; text-align: center; }}
                .summary-value {{ font-size: 24px; font-weight: bold; color: #10b981; }}
                .funnel {{ margin: 30px 0; }}
                .funnel-stage {{ display: flex; align-items: center; margin: 10px 0; }}
                .funnel-bar {{ height: 30px; background: linear-gradient(90deg, #10b981, #34d399); border-radius: 4px; margin-right: 10px; }}
                .funnel-label {{ min-width: 100px; }}
                .footer {{ margin-top: 40px; padding-top: 20px; border-top: 1px solid #eee; font-size: 12px; color: #666; }}
            </style>
        </head>
        <body>
            <h1>Relatório de Vendas</h1>
            <p>Período: {period}</p>
            
            <div class="summary">
                <div class="summary-card">
                    <div class="summary-value">{summary.get('total_leads', 0)}</div>
                    <div>Total de Leads</div>
                </div>
                <div class="summary-card">
                    <div class="summary-value">{summary.get('converted', 0)}</div>
                    <div>Convertidos</div>
                </div>
                <div class="summary-card">
                    <div class="summary-value">{summary.get('conversion_rate', 0) * 100:.1f}%</div>
                    <div>Taxa de Conversão</div>
                </div>
                <div class="summary-card">
                    <div class="summary-value">R$ {summary.get('total_value', 0):,.0f}</div>
                    <div>Valor Total</div>
                </div>
            </div>
            
            <h2>Funil de Vendas</h2>
            <div class="funnel">
                {''.join(f'''
                <div class="funnel-stage">
                    <div class="funnel-label">{stage.get("name")}</div>
                    <div class="funnel-bar" style="width: {min(stage.get("count", 0) * 2, 400)}px;"></div>
                    <div>{stage.get("count", 0)}</div>
                </div>
                ''' for stage in funnel.get('stages', []))}
            </div>
            
            <div class="footer">
                <p>Gerado em {datetime.now().strftime('%d/%m/%Y às %H:%M')} pelo BI Agent</p>
            </div>
        </body>
        </html>
        """
    
    def _render_marketing_html(self, data: Dict, period: str) -> str:
        """Renderiza HTML para relatório de marketing."""
        summary = data.get('summary', {})
        campaigns = data.get('campaigns', [])
        
        return f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="UTF-8">
            <title>Relatório de Marketing</title>
            <style>
                body {{ font-family: 'Segoe UI', Arial, sans-serif; padding: 40px; color: #333; }}
                h1 {{ color: #1a1a2e; border-bottom: 3px solid #8b5cf6; padding-bottom: 10px; }}
                .metrics {{ display: grid; grid-template-columns: repeat(5, 1fr); gap: 15px; margin: 30px 0; }}
                .metric {{ background: linear-gradient(135deg, #8b5cf6 0%, #a78bfa 100%); color: white; padding: 15px; border-radius: 8px; text-align: center; }}
                .metric-value {{ font-size: 24px; font-weight: bold; }}
                table {{ width: 100%; border-collapse: collapse; margin: 20px 0; }}
                th, td {{ padding: 12px; text-align: left; border-bottom: 1px solid #eee; }}
                th {{ background: #f8f9fa; font-weight: 600; }}
                .roas-good {{ color: #10b981; font-weight: bold; }}
                .roas-bad {{ color: #ef4444; font-weight: bold; }}
                .footer {{ margin-top: 40px; padding-top: 20px; border-top: 1px solid #eee; font-size: 12px; color: #666; }}
            </style>
        </head>
        <body>
            <h1>Relatório de Marketing</h1>
            <p>Período: {period}</p>
            
            <div class="metrics">
                <div class="metric">
                    <div class="metric-value">R$ {summary.get('total_spend', 0):,.0f}</div>
                    <div>Investimento</div>
                </div>
                <div class="metric">
                    <div class="metric-value">{summary.get('total_leads', 0)}</div>
                    <div>Leads</div>
                </div>
                <div class="metric">
                    <div class="metric-value">R$ {summary.get('cpl', 0):.2f}</div>
                    <div>CPL</div>
                </div>
                <div class="metric">
                    <div class="metric-value">{summary.get('roas', 0):.1f}x</div>
                    <div>ROAS</div>
                </div>
                <div class="metric">
                    <div class="metric-value">R$ {summary.get('cac', 0):.0f}</div>
                    <div>CAC</div>
                </div>
            </div>
            
            <h2>Campanhas</h2>
            <table>
                <thead>
                    <tr>
                        <th>Campanha</th>
                        <th>Investimento</th>
                        <th>Leads</th>
                        <th>CPL</th>
                        <th>ROAS</th>
                        <th>Status</th>
                    </tr>
                </thead>
                <tbody>
                    {''.join(f'''
                    <tr>
                        <td>{c.get("name")}</td>
                        <td>R$ {c.get("spend", 0):,.0f}</td>
                        <td>{c.get("leads", 0)}</td>
                        <td>R$ {c.get("cpl", 0):.2f}</td>
                        <td class="{'roas-good' if c.get('roas', 0) >= 1 else 'roas-bad'}">{c.get("roas", 0):.1f}x</td>
                        <td>{c.get("status", "active")}</td>
                    </tr>
                    ''' for c in campaigns)}
                </tbody>
            </table>
            
            <div class="footer">
                <p>Gerado em {datetime.now().strftime('%d/%m/%Y às %H:%M')} pelo BI Agent</p>
            </div>
        </body>
        </html>
        """
    
    async def _html_to_pdf(self, html: str, filepath: str) -> None:
        """Converte HTML para PDF usando WeasyPrint."""
        try:
            from weasyprint import HTML
            HTML(string=html).write_pdf(filepath)
        except ImportError:
            # Fallback: salva HTML se WeasyPrint não estiver instalado
            html_path = filepath.replace('.pdf', '.html')
            with open(html_path, 'w', encoding='utf-8') as f:
                f.write(html)
            logger.warning("WeasyPrint não instalado, HTML salvo em vez de PDF")


class ExcelExportService:
    """
    Serviço de exportação para Excel.
    
    Usa openpyxl para gerar planilhas.
    """
    
    def __init__(self, tenant_id: str):
        self.tenant_id = tenant_id
        self.output_dir = f"/tmp/exports/{tenant_id}"
        os.makedirs(self.output_dir, exist_ok=True)
    
    async def export_leads(
        self,
        data: List[Dict],
        period: str
    ) -> Dict[str, Any]:
        """Exporta dados de leads para Excel."""
        filename = f"leads_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join(self.output_dir, filename)
        
        await self._create_excel(data, filepath, "Leads")
        
        return {
            "filename": filename,
            "filepath": filepath,
            "url": f"/exports/{self.tenant_id}/{filename}",
            "rows": len(data),
            "size_kb": os.path.getsize(filepath) // 1024 if os.path.exists(filepath) else 0,
            "generated_at": datetime.now().isoformat(),
        }
    
    async def export_campaigns(
        self,
        data: List[Dict],
        period: str
    ) -> Dict[str, Any]:
        """Exporta dados de campanhas para Excel."""
        filename = f"campaigns_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join(self.output_dir, filename)
        
        await self._create_excel(data, filepath, "Campanhas")
        
        return {
            "filename": filename,
            "filepath": filepath,
            "url": f"/exports/{self.tenant_id}/{filename}",
            "rows": len(data),
            "size_kb": os.path.getsize(filepath) // 1024 if os.path.exists(filepath) else 0,
            "generated_at": datetime.now().isoformat(),
        }
    
    async def export_tickets(
        self,
        data: List[Dict],
        period: str
    ) -> Dict[str, Any]:
        """Exporta dados de tickets para Excel."""
        filename = f"tickets_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join(self.output_dir, filename)
        
        await self._create_excel(data, filepath, "Tickets")
        
        return {
            "filename": filename,
            "filepath": filepath,
            "url": f"/exports/{self.tenant_id}/{filename}",
            "rows": len(data),
            "size_kb": os.path.getsize(filepath) // 1024 if os.path.exists(filepath) else 0,
            "generated_at": datetime.now().isoformat(),
        }
    
    async def _create_excel(
        self,
        data: List[Dict],
        filepath: str,
        sheet_name: str
    ) -> None:
        """Cria arquivo Excel com os dados."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            
            wb = Workbook()
            ws = wb.active
            ws.title = sheet_name
            
            if not data:
                ws['A1'] = 'Sem dados para o período selecionado'
                wb.save(filepath)
                return
            
            # Headers
            headers = list(data[0].keys())
            header_fill = PatternFill(start_color="4361EE", end_color="4361EE", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header.replace('_', ' ').title())
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
            
            # Data
            for row_idx, row_data in enumerate(data, 2):
                for col_idx, header in enumerate(headers, 1):
                    value = row_data.get(header, '')
                    if isinstance(value, dict) or isinstance(value, list):
                        value = str(value)
                    ws.cell(row=row_idx, column=col_idx, value=value)
            
            # Auto-width columns
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column].width = adjusted_width
            
            wb.save(filepath)
            
        except ImportError:
            # Fallback: salva como CSV se openpyxl não estiver instalado
            import csv
            csv_path = filepath.replace('.xlsx', '.csv')
            with open(csv_path, 'w', newline='', encoding='utf-8') as f:
                if data:
                    writer = csv.DictWriter(f, fieldnames=data[0].keys())
                    writer.writeheader()
                    writer.writerows(data)
            logger.warning("openpyxl não instalado, CSV salvo em vez de Excel")

